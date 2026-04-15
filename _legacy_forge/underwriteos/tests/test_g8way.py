"""
Tests for the G8WAY read-only snapshot reader.

We don't ship a real filled G8WAY with the repo (PII + size), so these
tests build a minimal synthetic workbook that mirrors the sheet names
and label layout Keith uses in production.
"""
from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from underwriteos.extract.g8way import read_g8way_snapshot, G8waySnapshot


def _build_fake_g8way(path: Path) -> None:
    wb = openpyxl.Workbook()
    # Drop default sheet
    wb.remove(wb.active)

    # --- USDA BSE Calculation ---
    ws = wb.create_sheet("USDA BSE Calculation")
    ws["A2"] = "Borrower:"
    ws["B2"] = "Mirzai Group LLC"
    ws["A8"] = "Total Assets"
    ws["B8"] = 1_675_987
    ws["A9"] = "Less: Total Liabilities"
    ws["B9"] = 407_342
    ws["A10"] = "Net Worth (Equity)"
    ws["B10"] = 1_268_644

    # --- USDA Cash Flow ---
    ws = wb.create_sheet("USDA Cash Flow")
    ws["A2"] = "Borrower:"
    ws["B2"] = "Mirzai Group LLC"
    ws["A20"] = "DSCR Year 1"
    ws["B20"] = 1.23
    ws["A21"] = "DSCR Year 2"
    ws["B21"] = 1.25
    ws["A22"] = "Global DSCR"
    ws["B22"] = 1.10

    # --- USDA Debt Details ---
    ws = wb.create_sheet("USDA Debt Details")
    ws["B2"] = "Borrower:"
    ws["D2"] = "Mirzai Group LLC"
    ws["B5"] = "Loan Amount"
    ws["C5"] = 310_000
    ws["B6"] = "Term (Months)"
    ws["C6"] = 120
    ws["B7"] = "Rate"
    ws["C7"] = 0.1025

    wb.save(path)


def test_read_g8way_snapshot_happy_path(tmp_path: Path):
    fake = tmp_path / "fake_g8way.xlsx"
    _build_fake_g8way(fake)

    snap = read_g8way_snapshot(fake)

    assert isinstance(snap, G8waySnapshot)
    assert snap.borrower == "Mirzai Group LLC"
    assert snap.total_assets == 1_675_987
    assert snap.total_liabilities == 407_342
    assert snap.net_worth == 1_268_644
    assert snap.dscr_y1 == 1.23
    assert snap.dscr_y2 == 1.25
    assert snap.global_dscr == 1.10
    assert snap.requested_loan_amount == 310_000
    assert snap.requested_term_months == 120
    assert snap.requested_rate == pytest.approx(0.1025)


def test_read_g8way_snapshot_missing_file(tmp_path: Path):
    with pytest.raises(FileNotFoundError):
        read_g8way_snapshot(tmp_path / "nope.xlsx")


def test_dscr_uw_sheet_takes_priority(tmp_path: Path):
    """When DSCR (UW) sheet exists, it should be used instead of USDA Cash Flow."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # USDA Cash Flow has DSCR 1.23
    ws = wb.create_sheet("USDA Cash Flow")
    ws["A20"] = "DSCR Year 1"
    ws["B20"] = 1.23

    # DSCR (UW) has DSCR 1.655 — this is the authoritative sheet
    ws2 = wb.create_sheet("DSCR (UW)")
    ws2["A5"] = "Borrower DSCR"
    ws2["B5"] = 1.655
    ws2["A6"] = "Gross Receipts"
    ws2["B6"] = 188_300
    ws2["A10"] = "Annual Debt Service"
    ws2["B10"] = 50_530
    ws2["A12"] = "Cash Flow Available"
    ws2["B12"] = 83_640

    # Boarding Sheet for borrower name
    ws3 = wb.create_sheet("Boarding Sheet")
    ws3["A2"] = "Borrower:"
    ws3["B2"] = "Tropical Treasure Hunt Co"

    out = tmp_path / "priority_test.xlsx"
    wb.save(out)

    snap = read_g8way_snapshot(out)
    assert snap.sheet_used_for_dscr == "DSCR (UW)"
    assert snap.dscr_y1 == 1.655
    assert snap.gross_receipts_y1 == 188_300
    assert snap.annual_debt_service == 50_530
    assert snap.cash_flow_y1 == 83_640
    assert snap.borrower == "Tropical Treasure Hunt Co"


def test_program_sheet_override(tmp_path: Path):
    """Caller can pass primary_sheet to route to a program-specific sheet."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("CIF Cash Flow")
    ws["A5"] = "Borrower DSCR"
    ws["B5"] = 1.35
    out = tmp_path / "cif_override.xlsx"
    wb.save(out)

    snap = read_g8way_snapshot(out, primary_sheet="CIF Cash Flow")
    assert snap.sheet_used_for_dscr == "CIF Cash Flow"
    assert snap.dscr_y1 == 1.35


def test_read_g8way_snapshot_tolerates_missing_sheets(tmp_path: Path):
    # Workbook with ONLY the BSE sheet — cash flow / debt missing
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("USDA BSE Calculation")
    ws["A2"] = "Borrower:"
    ws["B2"] = "Minimal Co"
    ws["A8"] = "Total Assets"
    ws["B8"] = 100_000
    ws["A10"] = "Net Worth (Equity)"
    ws["B10"] = 50_000
    out = tmp_path / "minimal.xlsx"
    wb.save(out)

    snap = read_g8way_snapshot(out)
    assert snap.borrower == "Minimal Co"
    assert snap.total_assets == 100_000
    assert snap.net_worth == 50_000
    # Missing sheets → None, not an exception
    assert snap.dscr_y1 is None
    assert snap.requested_loan_amount is None
