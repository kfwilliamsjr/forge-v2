"""
G8WAY workbook reader — READ ONLY.

G8WAY is the Colony Bank / USDA B&I / SBA / CIF spreadsheet stack used for
borrower cash flow, pro forma balance sheet, debt details, and BSE
calculations. Keith's hard rule from CLAUDE.md:

    NEVER write to a G8WAY workbook programmatically. openpyxl corrupts
    these files — it drops embedded .wmf images and the Data Validation
    ranges. File corruption rate observed: ~18%. Manual entry is the
    only reliable method.

This module therefore only READS. It pulls borrower identity and a
handful of key financial anchor values out of a filled workbook so the
memo renderer and reconcile engine can cross-check extractor output
against what the underwriter actually typed into G8WAY.

Day 10.5 rewrite: program-aware sheet selection.

    primary_sheet    = "DSCR (UW)" — authoritative for all programs
    fallback_sheets  = ("Global DSCR (UW)", "Executive Summary",
                        "Boarding Sheet", "USDA Cash Flow", ...)

The caller can pass explicit sheet preferences from Program.primary_g8way_sheet
and Program.fallback_g8way_sheets. Defaults cover the common case.

Usage:
    from underwriteos.extract.g8way import read_g8way_snapshot
    snap = read_g8way_snapshot("/path/to/Deal_G8WAY.xlsx")
    snap.borrower          # "Mirzai Group LLC"
    snap.dscr_y1           # 1.23

Missing fields come back as None. Never raises on a malformed sheet —
the underwriter can fix the sheet and re-run.
"""
from __future__ import annotations

from dataclasses import dataclass, asdict
from pathlib import Path
from typing import Any, Optional, Sequence

import openpyxl  # type: ignore


@dataclass
class G8waySnapshot:
    source_path: str
    sheet_used_for_dscr: Optional[str] = None
    borrower: Optional[str] = None
    # Balance sheet (post-closing column preferred)
    total_assets: Optional[float] = None
    total_liabilities: Optional[float] = None
    net_worth: Optional[float] = None
    # Cash flow / DSCR
    dscr_y1: Optional[float] = None
    dscr_y2: Optional[float] = None
    global_dscr: Optional[float] = None
    # Income statement anchors from DSCR (UW) sheet
    gross_receipts_y1: Optional[float] = None
    gross_receipts_y2: Optional[float] = None
    net_income_y1: Optional[float] = None
    cash_flow_y1: Optional[float] = None
    annual_debt_service: Optional[float] = None
    # Debt
    requested_loan_amount: Optional[float] = None
    requested_rate: Optional[float] = None
    requested_term_months: Optional[int] = None

    def to_dict(self) -> dict:
        return asdict(self)


def _norm(v: Any) -> str:
    return str(v).strip().lower() if v is not None else ""


def _is_number(v: Any) -> bool:
    return isinstance(v, (int, float)) and not isinstance(v, bool)


def _find_label_value(
    ws, label_substring: str, *,
    max_row: int = 200,
    max_col: int = 30,
    value_col_offset_range: range = range(1, 8),
) -> Optional[float]:
    """
    Walk the sheet looking for a cell whose text contains `label_substring`
    (case-insensitive). When found, return the first numeric value in the
    cells immediately to the right within `value_col_offset_range`.
    """
    needle = label_substring.lower()
    for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
        for cell in row:
            if cell.value is None:
                continue
            if needle not in _norm(cell.value):
                continue
            for off in value_col_offset_range:
                try:
                    neighbor = ws.cell(row=cell.row, column=cell.column + off)
                except Exception:
                    continue
                if _is_number(neighbor.value):
                    return float(neighbor.value)
    return None


def _find_label_text(
    ws, label_substring: str, *,
    max_row: int = 20, max_col: int = 20,
) -> Optional[str]:
    needle = label_substring.lower()
    for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
        for cell in row:
            if cell.value is None or not isinstance(cell.value, str):
                continue
            if needle not in _norm(cell.value):
                continue
            for off in range(1, 10):
                try:
                    neighbor = ws.cell(row=cell.row, column=cell.column + off)
                except Exception:
                    continue
                if isinstance(neighbor.value, str) and neighbor.value.strip():
                    val = neighbor.value.strip()
                    if val.lower() in ("borrower name", "n/a", "tbd"):
                        continue
                    return val
    return None


def _first_available(wb, sheet_names: Sequence[str]) -> Optional[str]:
    """Return the first sheet name from the list that exists in the workbook."""
    for name in sheet_names:
        if name in wb.sheetnames:
            return name
    return None


# Default sheet search order for each data category.
_DSCR_SHEETS = ("DSCR (UW)", "Global DSCR (UW)", "Executive Summary",
                "USDA Cash Flow", "Boarding Sheet")
_BALANCE_SHEETS = ("PF Balance Sheet (Borrower)", "USDA BSE Calculation",
                   "USDA PF Balance Sheet", "Executive Summary")
_IDENTITY_SHEETS = ("Boarding Sheet", "Executive Summary", "USDA P&L",
                    "USDA BSE Calculation", "USDA Cash Flow", "DSCR (UW)")
_DEBT_SHEETS = ("Debt Schedule (Borrower)", "USDA Debt Details",
                "Boarding Sheet")


def read_g8way_snapshot(
    path: str | Path,
    primary_sheet: str = "DSCR (UW)",
    fallback_sheets: Sequence[str] = (),
) -> G8waySnapshot:
    """
    Read a filled G8WAY workbook and return a snapshot of the headline
    values. NEVER writes. Uses data_only=True so cached formula results
    are returned — the caller does not need Excel installed.

    Args:
        path: path to the G8WAY .xlsx file
        primary_sheet: preferred sheet for DSCR data (from Program.primary_g8way_sheet)
        fallback_sheets: ordered fallback sheets (from Program.fallback_g8way_sheets)
    """
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(str(p))

    wb = openpyxl.load_workbook(p, data_only=True, read_only=True)
    snap = G8waySnapshot(source_path=str(p))

    # Build the DSCR sheet search order: caller primary → caller fallbacks → defaults
    dscr_search = [primary_sheet] + list(fallback_sheets) + list(_DSCR_SHEETS)
    # Dedupe while preserving order
    seen: set[str] = set()
    dscr_search_deduped = []
    for s in dscr_search:
        if s not in seen:
            seen.add(s)
            dscr_search_deduped.append(s)

    # --- Borrower name ---
    for sheet in _IDENTITY_SHEETS:
        if sheet in wb.sheetnames:
            name = _find_label_text(wb[sheet], "borrower")
            if not name:
                name = _find_label_text(wb[sheet], "applicant")
            if name:
                snap.borrower = name
                break

    # --- Cash flow / DSCR --- (try sheets in priority order)
    dscr_sheet = _first_available(wb, dscr_search_deduped)
    if dscr_sheet:
        snap.sheet_used_for_dscr = dscr_sheet
        ws = wb[dscr_sheet]

        # DSCR values — try multiple label variations
        for label in ("dscr year 1", "dscr y1", "borrower dscr", "dscr"):
            v = _find_label_value(ws, label)
            if v is not None:
                snap.dscr_y1 = v
                break
        for label in ("dscr year 2", "dscr y2"):
            v = _find_label_value(ws, label)
            if v is not None:
                snap.dscr_y2 = v
                break
        for label in ("global dscr",):
            snap.global_dscr = _find_label_value(ws, label)

        # Income statement anchors
        snap.gross_receipts_y1 = (
            _find_label_value(ws, "gross receipts") or
            _find_label_value(ws, "total revenue") or
            _find_label_value(ws, "sales")
        )
        snap.net_income_y1 = (
            _find_label_value(ws, "net income") or
            _find_label_value(ws, "net profit")
        )
        snap.cash_flow_y1 = (
            _find_label_value(ws, "cash flow available") or
            _find_label_value(ws, "available cash flow") or
            _find_label_value(ws, "total cash flow")
        )
        snap.annual_debt_service = (
            _find_label_value(ws, "annual debt service") or
            _find_label_value(ws, "total debt service") or
            _find_label_value(ws, "ads")
        )

    # If primary DSCR sheet had no DSCR values, try fallbacks
    if snap.dscr_y1 is None:
        for sheet_name in dscr_search_deduped:
            if sheet_name == dscr_sheet or sheet_name not in wb.sheetnames:
                continue
            ws2 = wb[sheet_name]
            for label in ("dscr year 1", "dscr y1", "borrower dscr", "dscr"):
                v = _find_label_value(ws2, label)
                if v is not None:
                    snap.dscr_y1 = v
                    snap.sheet_used_for_dscr = sheet_name
                    break
            if snap.dscr_y1 is not None:
                break

    # --- Balance sheet ---
    bs_sheet = _first_available(wb, _BALANCE_SHEETS)
    if bs_sheet:
        ws = wb[bs_sheet]
        snap.total_assets = _find_label_value(ws, "total assets", max_row=60)
        snap.total_liabilities = _find_label_value(ws, "total liabilities", max_row=60)
        snap.net_worth = _find_label_value(ws, "net worth", max_row=60)

    # --- Debt details ---
    debt_sheet = _first_available(wb, _DEBT_SHEETS)
    if debt_sheet:
        ws = wb[debt_sheet]
        snap.requested_loan_amount = _find_label_value(ws, "loan amount", max_row=30)
        snap.requested_rate = _find_label_value(ws, "rate", max_row=30)
        term = _find_label_value(ws, "term (months)", max_row=30)
        if term is None:
            term = _find_label_value(ws, "term", max_row=30)
        snap.requested_term_months = int(term) if term is not None else None

    return snap
